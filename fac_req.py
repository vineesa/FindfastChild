import face_recognition
import pickle
import datetime
import cv2
import os
import pandas as pd
from openpyxl import Workbook, load_workbook

# Initialize 'currentname' to track new identifications
currentname = "unknown"

# Load face encodings
encodingsP = "encodings.pickle"
image_path = "C:\\Users\\bhara\\Downloads\\project\\Bharat\\dataset\\testing\\image_0.jpg"  # Input image path

print("[INFO] Loading encodings...")
data = pickle.loads(open(encodingsP, "rb").read())

# Load the image
image = cv2.imread(image_path)
if image is None:
    print("[ERROR] Image not found. Please check the path.")
    exit()

# Convert image to RGB (face_recognition requires RGB format)
rgb_image = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)

# Detect faces and compute embeddings
boxes = face_recognition.face_locations(rgb_image)
encodings = face_recognition.face_encodings(rgb_image, boxes)
names = []

# Prepare Excel file for recognized names
excel_file = "recognized_faces.xlsx"
if not os.path.exists(excel_file):
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Recognitions"
    sheet.append(["Timestamp", "Name"])
    wb.save(excel_file)

# Load existing Excel file
wb = load_workbook(excel_file)
sheet = wb["Recognitions"]

# Load Missing Children Details from Excel
details_file = "missing_children.xlsx"  # Ensure this file contains Name, Age, Gender, Missing Place
if os.path.exists(details_file):
    df = pd.read_excel(details_file)
else:
    print("[ERROR] Missing children details file not found.")
    df = pd.DataFrame(columns=["Name", "Age", "Gender", "Missing Place"])  # Create empty DataFrame

# Face recognition and detail retrieval
for encoding in encodings:
    matches = face_recognition.compare_faces(data["encodings"], encoding)
    name = "Unknown"
    
    if True in matches:
        matchedIdxs = [i for (i, b) in enumerate(matches) if b]
        counts = {}
        
        for i in matchedIdxs:
            name = data["names"][i]
            counts[name] = counts.get(name, 0) + 1
        
        name = max(counts, key=counts.get)

        if currentname != name:
            currentname = name
            print(f"🔹 Recognized: {currentname}")

            # ✅ Retrieve person details from Excel
            person_details = df[df["Name"].str.lower() == currentname.lower()]
            if not person_details.empty:
                details = person_details.iloc[0].to_dict()
                print(f"✅ Details Found: {details}")
            else:
                print("⚠ No additional details found in the database.")

            # ✅ Log the recognition in the Excel sheet
            timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            sheet.append([timestamp, currentname])
            wb.save(excel_file)

    names.append(name)

# Draw bounding boxes and labels
for ((top, right, bottom, left), name) in zip(boxes, names):
    cv2.rectangle(image, (left, top), (right, bottom), (0, 255, 0), 2)
    y = top - 15 if top - 15 > 15 else top + 15
    cv2.putText(image, name, (left, y), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 255, 0), 2)

# Save the output image
output_path = "output.jpg"
cv2.imwrite(output_path, image)
print("[INFO] Output image saved at:", output_path)

# Cleanup
wb.save(excel_file)
