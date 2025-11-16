import os
import cv2
import pickle
import datetime
import face_recognition
import pandas as pd
from flask import Flask, render_template, request, redirect, url_for, jsonify, session
from openpyxl import Workbook, load_workbook
import imaplib
import email
import smtplib
from email.header import decode_header
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage


# Initialize Flask app
app = Flask(__name__)
app.secret_key = "your_secret_key"  # Secret key for session handling

# Define image upload folder
UPLOAD_FOLDER = "static/uploads"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Load face encodings
encodingsP = "encodings.pickle"
if not os.path.exists(encodingsP):
    print("[ERROR] Encodings file not found. Make sure 'encodings.pickle' exists.")
    exit()

print("[INFO] Loading face encodings...")
data = pickle.loads(open(encodingsP, "rb").read())

# Prepare Excel file for logging recognized names
excel_file = "recognized_faces.xlsx"
if not os.path.exists(excel_file):
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Recognitions"
    sheet.append(["Timestamp", "Name"])
    wb.save(excel_file)

# Load existing Excel file
wb = load_workbook(excel_file)
sheet = wb["Recognitions"]

# Email Configuration
EMAIL = "johndeofficial333@gmail.com"  # Replace with your email
PASSWORD = "jvxq ahxb iheu pjju"  # Replace with your email password
SMTP_SERVER = "smtp.gmail.com"
IMAP_SERVER = "imap.gmail.com"
SMTP_PORT = 587
IMAP_PORT = 993
EMAIL2 = "swethanadh17@gmail.com"

# Function to send an email with an image attachment
def send_email_with_image(image_path):
    try:
        msg = MIMEMultipart()
        msg['From'] = EMAIL
        msg['To'] = EMAIL2
        msg['Subject'] = "Person Detected"

        body = "An known person was detected. See the attached image."
        msg.attach(MIMEText(body, 'plain'))

        with open(image_path, 'rb') as f:
            img_data = f.read()
            image = MIMEImage(img_data, name=os.path.basename(image_path))
            msg.attach(image)

        server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
        server.starttls()
        server.login(EMAIL, PASSWORD)
        server.sendmail(EMAIL, EMAIL2, msg.as_string())
        server.quit()
        print("Email sent with image attachment.")

    except Exception as e:
        print(f"Error sending email: {e}")



# Load Missing Children Details from Excel
details_file = "missing_children.xlsx"

def load_data():
    """Loads missing children details from the Excel file."""
    if os.path.exists(details_file):
        return pd.read_excel(details_file)
    else:
        print("[ERROR] Missing children details file not found.")
        return pd.DataFrame(columns=["Name", "Age", "Gender", "Missing Place"])

def save_data(df):
    """Saves the updated missing children data back to the Excel file."""
    df.to_excel(details_file, index=False)

df = load_data()

# 🔹 Face Recognition Function
def recognize_faces(image_path):
    image = cv2.imread(image_path)
    if image is None:
        return [], None, {}

    rgb_image = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
    boxes = face_recognition.face_locations(rgb_image)
    encodings = face_recognition.face_encodings(rgb_image, boxes)
    names = []
    details = {}

    for encoding in encodings:
        matches = face_recognition.compare_faces(data["encodings"], encoding)
        name = "Unknown"

        if True in matches:
            matchedIdxs = [i for (i, b) in enumerate(matches) if b]
            counts = {}

            for i in matchedIdxs:
                name = data["names"][i]
                counts[name] = counts.get(name, 0) + 1

            name = max(counts, key=counts.get)

            # Retrieve person's details from Excel
            person_details = df[df["Name"].str.lower() == name.lower()]
            if not person_details.empty:
                details = person_details.iloc[0].to_dict()
                # Save detected known person's image with timestamp
                timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
                known_image_path = os.path.join(
                    app.config['UPLOAD_FOLDER'], 
                    f"detected_{name}_{timestamp}.jpg"
                )
                cv2.imwrite(known_image_path, image)
                print(f"[INFO] Known person detected: {name}")
                send_email_with_image(known_image_path)

            # Log recognized face in Excel
            timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            sheet.append([timestamp, name])
            wb.save(excel_file)

        names.append(name)
    '''image_path = "unknown_person.jpg"
    cv2.imwrite(image_path, frame)
    print("Unknown person detected. Image captured.")
    send_email_with_image(image_path)'''

    # Draw bounding boxes and labels
    for ((top, right, bottom, left), name) in zip(boxes, names):
        cv2.rectangle(image, (left, top), (right, bottom), (0, 255, 0), 2)
        y = top - 15 if top - 15 > 15 else top + 15
        cv2.putText(image, name, (left, y), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 255, 0), 2)

    output_path = os.path.join(app.config['UPLOAD_FOLDER'], "output.jpg")
    cv2.imwrite(output_path, image)

    return names, output_path, details

# 🔹 Flask Routes
@app.route("/")
def login_page():
    """Login Page Route"""
    return render_template("page1.html")

@app.route("/home", methods=["GET", "POST"])
def home_page():
    """User Home Page Route"""
    if request.method == "POST":
        email = request.form.get("email")
        password = request.form.get("password")
        if email and password:  # Authentication logic can be added here
            session["user"] = email  # Store user session
            return redirect(url_for("home_page"))
    return render_template("home.html")

# ✅ Admin Login & Dashboard
@app.route("/admin_login", methods=["GET", "POST"])
def admin_login():
    """Handles Admin Login"""
    if "admin" in session:
        return redirect(url_for("admin_dashboard"))

    if request.method == "POST":
        email = request.form.get("email")
        password = request.form.get("password")

        # Sample authentication logic (Replace with a database check)
        if email == "admin@example.com" and password == "admin123":
            session["admin"] = email
            return redirect(url_for("admin_dashboard"))
        else:
            return render_template("page1.html", error="Invalid admin credentials!", active_section="admin-auth")

    return redirect(url_for("login_page"))

@app.route("/admin_dashboard")
def admin_dashboard():
    """Admin Dashboard Route"""
    if "admin" not in session:
        return redirect(url_for("login_page"))  # Redirect if not logged in

    df = load_data()
    cases = df.to_dict(orient="records")  # Convert DataFrame to list of dictionaries
    return render_template("admin.html", cases=cases)

# ✅ Edit Case Route
@app.route("/edit_case", methods=["POST"])
def edit_case():
    """Edits a missing child record"""
    data = request.json
    df = load_data()

    # Find and update the record
    index = df[df["Name"] == data["old_name"]].index
    if not index.empty:
        df.loc[index, ["Name", "Age", "Gender", "Missing Place"]] = [data["name"], data["age"], data["gender"], data["missing_place"]]
        save_data(df)
        return jsonify({"message": "Case updated successfully!"})
    else:
        return jsonify({"error": "Record not found!"}), 400

# ✅ Delete Case Route
@app.route("/delete_case", methods=["POST"])
def delete_case():
    """Deletes a missing child record"""
    data = request.json
    df = load_data()

    df = df[df["Name"] != data["name"]]  # Remove record
    save_data(df)
    return jsonify({"message": "Case deleted successfully!"})

@app.route("/add_case", methods=["POST"])
def add_case():
    """Adds a new missing child record"""
    if "admin" not in session:
        return jsonify({"error": "Unauthorized"}), 401

    data = request.json
    df = load_data()
    
    # Create new record
    new_record = {
        "Name": data["name"],
        "Age": data["age"],
        "Gender": data["gender"],
        "Missing Place": data["missing_place"],
        "Contact No":data["contact_no"],
        "Email":data["email"],
    }
    
    # Append new record to DataFrame
    df = pd.concat([df, pd.DataFrame([new_record])], ignore_index=True)
    save_data(df)
    
    return jsonify({"message": "Case added successfully!"})

@app.route("/upload", methods=["POST"])
def upload():
    """Handles Image Upload for Face Recognition"""
    if "child-photo" not in request.files:
        return render_template("home.html", result="No file uploaded")

    file = request.files["child-photo"]
    if file.filename == "":
        return render_template("home.html", result="No file selected")

    file_path = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
    file.save(file_path)

    recognized_names, output_image, details = recognize_faces(file_path)
    result_msg = f"Recognized: {', '.join(recognized_names)}" if recognized_names else "No match found"

    return render_template("home.html", result=result_msg, file_path=output_image, details=details)

@app.route("/logout")
def logout():
    """Logs out the user/admin"""
    session.pop("user", None)
    session.pop("admin", None)
    return redirect(url_for("login_page"))

if __name__ == "__main__":
    app.run(debug=True)
